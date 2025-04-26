# api_final.py
from flask import Flask, request, jsonify, render_template, Response, url_for
import os

# Force Matplotlib to use a writable cache directory
os.environ['MPLCONFIGDIR'] = '/tmp/.matplotlib'

from openpyxl import load_workbook
from astropy.coordinates import SkyCoord, EarthLocation
from astropy.time import Time
import astropy.units as u
from astroplan import Observer, FixedTarget
from datetime import datetime, timedelta
from pytz import timezone
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from io import BytesIO

app = Flask(__name__)

# مسار ملف الإكسل
script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, 'Objects.xlsx')
if not os.path.exists(file_path):
    raise FileNotFoundError(f"File not found: {file_path}")

# قراءة ورقة العمل
wb = load_workbook(file_path, data_only=True)
ws = wb.active

# استخراج أسماء الأعمدة لبناء فهارس
header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
idx = {
    'object': header.index('object'),
    'R.A._H': header.index('R.A._H'),
    'R.A._M': header.index('R.A._M'),
    'R.A._S': header.index('R.A._S'),
    'DEC._H': header.index('DEC._H'),
    'DEC._M': header.index('DEC._M'),
    'DEC._S': header.index('DEC._S'),
}

# تحميل البيانات في قائمة من القواميس
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    obj_name = row[idx['object']]
    data.append({
        'object': str(obj_name).strip().lower(),
        'R.A._H': float(row[idx['R.A._H']]),
        'R.A._M': float(row[idx['R.A._M']]),
        'R.A._S': float(row[idx['R.A._S']]),
        'DEC._H': float(row[idx['DEC._H']]),
        'DEC._M': float(row[idx['DEC._M']]),
        'DEC._S': float(row[idx['DEC._S']]),
    })

# إعداد موقع الراصد
latitude = 33.27427886628448
longitude = 44.3800838290597
elevation = 40
location = EarthLocation(
    lat=latitude * u.deg,
    lon=longitude * u.deg,
    height=elevation * u.m
)

local_tz = timezone('Asia/Baghdad')
observer = Observer(location=location, timezone=local_tz)

def get_target(object_name: str) -> FixedTarget:
    key = object_name.strip().lower()
    matches = [d for d in data if d['object'] == key]
    if not matches:
        return None
    obj = matches[0]
    ra = obj['R.A._H'] + obj['R.A._M'] / 60 + obj['R.A._S'] / 3600
    dec = obj['DEC._H'] + obj['DEC._M'] / 60 + obj['DEC._S'] / 3600
    coord = SkyCoord(ra=ra * u.hour, dec=dec * u.deg, frame='icrs')
    return FixedTarget(coord=coord, name=object_name)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/plot', methods=['GET'])
def plot_altitude():
    import matplotlib.dates as mdates

    object_name = request.args.get('object', '').strip()
    date_str = request.args.get('date', '')
    if not object_name:
        return jsonify({"error": "The 'object' parameter is required."}), 400
    if not date_str:
        return jsonify({"error": "The 'date' parameter is required (format: YYYY-MM-DD)."}), 400

    try:
        date_local = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "Invalid date format. Use 'YYYY-MM-DD'."}), 400

    target = get_target(object_name)
    if target is None:
        return jsonify({"error": f"Object '{object_name}' not found."}), 404

    times = [local_tz.localize(date_local + timedelta(minutes=15 * i)) for i in range(96)]
    mdates_list, altitudes = [], []

    for t in times:
        t_astropy = Time(t)
        altaz = observer.altaz(t_astropy, target)
        mdates_list.append(mdates.date2num(t))
        altitudes.append(altaz.alt.degree)

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(mdates_list, altitudes, 'o-', label='Altitude')
    ax.set_title(f"Altitude vs Time for {object_name} on {date_str}")
    ax.set_xlabel('Local Time')
    ax.set_ylabel('Altitude (deg)')
    ax.grid(True)
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
    fig.autofmt_xdate()

    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)

    return Response(buf.getvalue(), mimetype='image/png')

@app.route('/plot_page', methods=['GET'])
def plot_page():
    from matplotlib import dates as mdates

    dt_str = request.args.get('datetime', '').strip()
    object_name = request.args.get('object', '').strip()
    if not dt_str or not object_name:
        return "Missing 'datetime' or 'object' parameter.", 400

    try:
        dt_naive = datetime.fromisoformat(dt_str)
    except ValueError:
        return "Invalid datetime format. Use 'YYYY-MM-DDTHH:MM' or 'YYYY-MM-DDTHH:MM:SS'.", 400
    dt_local = local_tz.localize(dt_naive)

    target = get_target(object_name)
    if target is None:
        return f"Object '{object_name}' not found.", 404

    altaz = observer.altaz(Time(dt_local), target)
    altitude = altaz.alt.degree
    azimuth = altaz.az.degree

    sunrise_time = observer.target_rise_time(Time(dt_local), target, which='next')
    sunset_time  = observer.target_set_time(Time(dt_local), target, which='next')

    def format_time(t):
        if getattr(t, 'mask', False):
            return "Always visible"
        return t.to_datetime(timezone=local_tz).strftime('%H:%M')

    sunrise = format_time(sunrise_time)
    sunset  = format_time(sunset_time)

    plot_url = url_for('plot_altitude', object=object_name,
                       date=dt_local.strftime("%Y-%m-%d"))

    return render_template('plot_page.html',
                           plot_url=plot_url,
                           object_name=object_name,
                           datetime_str=dt_local.strftime("%Y-%m-%d %H:%M"),
                           altitude=round(altitude, 3),
                           azimuth=round(azimuth, 3),
                           rise=sunrise,
                           set=sunset)

if __name__ == '__main__':
    app.run(debug=True)
